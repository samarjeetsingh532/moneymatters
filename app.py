import calendar
import io
import os
import sqlite3
from datetime import date, datetime

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from werkzeug.security import check_password_hash

from database.db import (
    create_user,
    ensure_admin_user,
    get_user_by_email,
    init_db,
    seed_db,
)
from database.queries import (
    CURRENCIES,
    currency_symbol,
    delete_account_by_id,
    delete_expense_by_id,
    delete_income_by_id,
    delete_user_by_id,
    get_account_breakdown,
    get_account_by_id,
    get_accounts_by_user,
    get_all_transactions,
    get_all_users,
    get_category_breakdown,
    get_default_account,
    get_expense_by_id,
    get_income_by_id,
    get_recent_transactions,
    get_summary_stats,
    get_user_by_id,
    get_user_for_admin,
    insert_account,
    insert_expense,
    insert_income,
    update_account,
    update_expense,
    update_income,
    update_user_base_currency,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")

CATEGORIES = [
    "Food",
    "Transport",
    "Bills",
    "Health",
    "Entertainment",
    "Shopping",
    "Other",
]

INCOME_SOURCES = [
    "Salary",
    "Freelance",
    "Business",
    "Investment",
    "Gift",
    "Other",
]

ACCOUNT_TYPES = [
    "Bank Account",
    "Cash",
    "Credit Card",
    "Wallet",
    "Other",
]

with app.app_context():
    init_db()
    seed_db()
    ensure_admin_user()


def _parse_date(val):
    try:
        datetime.strptime(val, "%Y-%m-%d")
        return val
    except (ValueError, TypeError):
        return None


def _months_ago(today, n):
    m, y = today.month - n, today.year
    while m <= 0:
        m += 12
        y -= 1
    return date(y, m, 1).isoformat()


def _resolve_selected_account(user_id, raw_account_id):
    """Validate the submitted account belongs to the user and is active;
    fall back to their default account otherwise (e.g. old forms/tests that
    don't submit an account at all)."""
    account_id = None
    try:
        candidate_id = int(raw_account_id)
    except (TypeError, ValueError):
        candidate_id = None

    if candidate_id is not None:
        account = get_account_by_id(candidate_id, user_id)
        if account is not None and account["is_active"]:
            account_id = candidate_id

    if account_id is None:
        default_account = get_default_account(user_id)
        account_id = default_account["id"] if default_account else None

    return account_id


def _parse_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _build_transactions_workbook(transactions):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"

    sheet.append(["Date", "Description", "Category", "Type", "Account", "Currency", "Amount"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for tx in transactions:
        sheet.append(
            [
                datetime.strptime(tx["date"], "%Y-%m-%d").date(),
                tx["description"],
                tx["category"],
                "Income" if tx["type"] == "income" else "Expense",
                tx["account_name"],
                tx["currency"],
                tx["amount"],
            ]
        )

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "yyyy-mm-dd"
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
        row[0].number_format = "#,##0.00"

    widths = {"A": 14, "B": 34, "C": 16, "D": 10, "E": 18, "F": 10, "G": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    return workbook


def _send_workbook(workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ------------------------------------------------------------------ #
# Routes                                                              #
# ------------------------------------------------------------------ #


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if session.get("user_id"):
        return redirect(url_for("profile"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not all([name, email, password, confirm_password]):
            flash("All fields are required.", "error")
            return render_template("register.html")

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("register.html")

        try:
            create_user(name, email, password)
        except sqlite3.IntegrityError:
            flash("Email already registered.", "error")
            return render_template("register.html")

        flash("Account created! Please sign in.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("profile"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        user = get_user_by_email(email)
        if not user or not check_password_hash(user["password_hash"], password):
            flash("Invalid email or password.", "error")
            return render_template("login.html")

        session["user_id"] = user["id"]
        session["user_name"] = user["name"]
        session["is_admin"] = bool(user["is_admin"])
        return redirect(url_for("profile"))

    return render_template("login.html")


@app.route("/admin")
def admin_users():
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if not session.get("is_admin"):
        abort(403)
    return render_template("admin_users.html", users=get_all_users())


@app.route("/admin/users/<int:user_id>")
def admin_user_detail(user_id):
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if not session.get("is_admin"):
        abort(403)

    target_user = get_user_for_admin(user_id)
    if target_user is None:
        abort(404)

    return render_template(
        "admin_user_detail.html",
        target_user=target_user,
        accounts=get_accounts_by_user(user_id),
        transactions=get_all_transactions(user_id),
        currency_symbol=currency_symbol,
    )


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
def admin_delete_user(user_id):
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if not session.get("is_admin"):
        abort(403)

    if user_id == session.get("user_id"):
        flash("You cannot delete your own account.", "error")
        return redirect(url_for("admin_users"))

    target_user = get_user_for_admin(user_id)
    if target_user is None:
        abort(404)

    delete_user_by_id(user_id)
    flash(f"Deleted {target_user['name']} and all their data.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/accounts/<int:account_id>/delete", methods=["POST"])
def admin_delete_account(user_id, account_id):
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if not session.get("is_admin"):
        abort(403)

    if get_user_for_admin(user_id) is None:
        abort(404)

    delete_account_by_id(account_id, user_id)
    flash("Account removed.", "success")
    return redirect(url_for("admin_user_detail", user_id=user_id))


# ------------------------------------------------------------------ #
# Placeholder routes — students will implement these                  #
# ------------------------------------------------------------------ #


@app.route("/terms")
def terms():
    return render_template("terms.html")


@app.route("/privacy")
def privacy():
    return render_template("privacy.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("landing"))


@app.route("/profile")
def profile():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    uid = session["user_id"]
    today = date.today()
    user = get_user_by_id(uid)

    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))
    account_id = _parse_int(request.args.get("account"))
    currency_filter = request.args.get("currency") or None
    if currency_filter not in CURRENCIES:
        currency_filter = None
    tx_type = request.args.get("type") or None
    if tx_type not in ("income", "expense"):
        tx_type = None

    if date_from and date_to and date_from > date_to:
        flash("Start date must be before end date.", "error")
        date_from = date_to = None

    today_str = today.isoformat()
    this_month_from = today.replace(day=1).isoformat()
    this_month_to = today.replace(
        day=calendar.monthrange(today.year, today.month)[1]
    ).isoformat()

    presets = {
        "this_month": {"date_from": this_month_from, "date_to": this_month_to},
        "last_3": {"date_from": _months_ago(today, 3), "date_to": today_str},
        "last_6": {"date_from": _months_ago(today, 6), "date_to": today_str},
    }

    base_currency = user["base_currency"] if user else "INR"

    return render_template(
        "profile.html",
        user=user,
        stats=get_summary_stats(
            uid,
            date_from,
            date_to,
            base_currency=base_currency,
            account_id=account_id,
            tx_type=tx_type,
            currency=currency_filter,
        ),
        transactions=get_recent_transactions(
            uid,
            date_from=date_from,
            date_to=date_to,
            account_id=account_id,
            tx_type=tx_type,
            currency=currency_filter,
        ),
        categories=get_category_breakdown(
            uid,
            date_from,
            date_to,
            base_currency=base_currency,
            account_id=account_id,
            currency=currency_filter,
        ),
        accounts=get_accounts_by_user(uid),
        account_breakdown=get_account_breakdown(uid, date_from, date_to),
        currency_symbol=currency_symbol,
        currencies=CURRENCIES,
        date_from=date_from,
        date_to=date_to,
        account_id=account_id,
        currency_filter=currency_filter,
        tx_type=tx_type,
        presets=presets,
        month_names=list(calendar.month_name)[1:],
        export_years=list(range(today.year - 5, today.year + 1)),
        current_month=today.month,
        current_year=today.year,
    )


@app.route("/base-currency", methods=["POST"])
def update_base_currency():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    currency = request.form.get("base_currency", "").strip()
    if currency in CURRENCIES:
        update_user_base_currency(session["user_id"], currency)
        flash("Base currency updated.", "success")
    else:
        flash("Please select a valid currency.", "error")

    return redirect(url_for("profile"))


@app.route("/expenses/add", methods=["GET", "POST"])
def add_expense():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    today = date.today().isoformat()
    accounts = get_accounts_by_user(session["user_id"], active_only=True)

    if request.method == "POST":
        amount_raw = request.form.get("amount", "").strip()
        category = request.form.get("category", "").strip()
        expense_date = request.form.get("date", "").strip()
        description = request.form.get("description", "").strip()
        account_id = _resolve_selected_account(
            session["user_id"], request.form.get("account")
        )

        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except ValueError:
            flash("Amount must be a positive number.", "error")
            return render_template(
                "add_expense.html",
                categories=CATEGORIES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        if category not in CATEGORIES:
            flash("Please select a valid category.", "error")
            return render_template(
                "add_expense.html",
                categories=CATEGORIES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        if not _parse_date(expense_date):
            flash("Please enter a valid date.", "error")
            return render_template(
                "add_expense.html",
                categories=CATEGORIES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        insert_expense(
            session["user_id"], amount, category, expense_date, description,
            account_id=account_id,
        )
        flash("Expense added.", "success")
        return redirect(url_for("profile"))

    return render_template(
        "add_expense.html", categories=CATEGORIES, accounts=accounts, currencies=CURRENCIES, form={}, today=today
    )


@app.route("/expenses/<int:id>/edit", methods=["GET", "POST"])
def edit_expense(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    expense = get_expense_by_id(id, session["user_id"])
    if expense is None:
        abort(404)

    accounts = get_accounts_by_user(session["user_id"], active_only=True)

    if request.method == "GET":
        return render_template(
            "edit_expense.html",
            expense=expense,
            categories=CATEGORIES,
            accounts=accounts, currencies=CURRENCIES,
            form={},
        )

    amount_raw = request.form.get("amount", "").strip()
    category = request.form.get("category", "").strip()
    expense_date = request.form.get("date", "").strip()
    description = request.form.get("description", "").strip()
    account_id = _resolve_selected_account(session["user_id"], request.form.get("account"))

    try:
        amount = float(amount_raw)
        if amount <= 0:
            raise ValueError
    except ValueError:
        flash("Amount must be a positive number.", "error")
        return render_template(
            "edit_expense.html",
            expense=expense,
            categories=CATEGORIES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    if category not in CATEGORIES:
        flash("Please select a valid category.", "error")
        return render_template(
            "edit_expense.html",
            expense=expense,
            categories=CATEGORIES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    if not _parse_date(expense_date):
        flash("Please enter a valid date.", "error")
        return render_template(
            "edit_expense.html",
            expense=expense,
            categories=CATEGORIES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    update_expense(
        id, session["user_id"], amount, category, expense_date, description,
        account_id=account_id,
    )
    flash("Expense updated.", "success")
    return redirect(url_for("profile"))


@app.route("/expenses/<int:id>/delete", methods=["POST"])
def delete_expense(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    expense = get_expense_by_id(id, session["user_id"])
    if expense is None:
        abort(404)

    delete_expense_by_id(id, session["user_id"])
    return redirect(url_for("profile"))


@app.route("/income/add", methods=["GET", "POST"])
def add_income():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    today = date.today().isoformat()
    accounts = get_accounts_by_user(session["user_id"], active_only=True)

    if request.method == "POST":
        amount_raw = request.form.get("amount", "").strip()
        source = request.form.get("source", "").strip()
        income_date = request.form.get("date", "").strip()
        description = request.form.get("description", "").strip()
        account_id = _resolve_selected_account(
            session["user_id"], request.form.get("account")
        )

        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except ValueError:
            flash("Amount must be a positive number.", "error")
            return render_template(
                "add_income.html",
                sources=INCOME_SOURCES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        if source not in INCOME_SOURCES:
            flash("Please select a valid source.", "error")
            return render_template(
                "add_income.html",
                sources=INCOME_SOURCES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        if not _parse_date(income_date):
            flash("Please enter a valid date.", "error")
            return render_template(
                "add_income.html",
                sources=INCOME_SOURCES,
                accounts=accounts, currencies=CURRENCIES,
                form=request.form,
                today=today,
            )

        insert_income(
            session["user_id"], amount, source, income_date, description,
            account_id=account_id,
        )
        flash("Income added.", "success")
        return redirect(url_for("profile"))

    return render_template(
        "add_income.html", sources=INCOME_SOURCES, accounts=accounts, currencies=CURRENCIES, form={}, today=today
    )


@app.route("/income/<int:id>/edit", methods=["GET", "POST"])
def edit_income(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    income = get_income_by_id(id, session["user_id"])
    if income is None:
        abort(404)

    accounts = get_accounts_by_user(session["user_id"], active_only=True)

    if request.method == "GET":
        return render_template(
            "edit_income.html",
            income=income,
            sources=INCOME_SOURCES,
            accounts=accounts, currencies=CURRENCIES,
            form={},
        )

    amount_raw = request.form.get("amount", "").strip()
    source = request.form.get("source", "").strip()
    income_date = request.form.get("date", "").strip()
    description = request.form.get("description", "").strip()
    account_id = _resolve_selected_account(session["user_id"], request.form.get("account"))

    try:
        amount = float(amount_raw)
        if amount <= 0:
            raise ValueError
    except ValueError:
        flash("Amount must be a positive number.", "error")
        return render_template(
            "edit_income.html",
            income=income,
            sources=INCOME_SOURCES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    if source not in INCOME_SOURCES:
        flash("Please select a valid source.", "error")
        return render_template(
            "edit_income.html",
            income=income,
            sources=INCOME_SOURCES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    if not _parse_date(income_date):
        flash("Please enter a valid date.", "error")
        return render_template(
            "edit_income.html",
            income=income,
            sources=INCOME_SOURCES,
            accounts=accounts, currencies=CURRENCIES,
            form=request.form,
        )

    update_income(
        id, session["user_id"], amount, source, income_date, description,
        account_id=account_id,
    )
    flash("Income updated.", "success")
    return redirect(url_for("profile"))


@app.route("/income/<int:id>/delete", methods=["POST"])
def delete_income(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    income = get_income_by_id(id, session["user_id"])
    if income is None:
        abort(404)

    delete_income_by_id(id, session["user_id"])
    return redirect(url_for("profile"))


@app.route("/accounts")
def accounts():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    return render_template(
        "accounts.html",
        accounts=get_accounts_by_user(session["user_id"]),
        currency_symbol=currency_symbol,
    )


@app.route("/accounts/add", methods=["GET", "POST"])
def add_account():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        account_type = request.form.get("account_type", "").strip()
        currency = request.form.get("currency", "").strip()
        opening_balance_raw = request.form.get("opening_balance", "").strip()
        description = request.form.get("description", "").strip()

        if not name:
            flash("Please enter an account name.", "error")
            return render_template(
                "add_account.html", account_types=ACCOUNT_TYPES, currencies=CURRENCIES,
                form=request.form,
            )

        if account_type not in ACCOUNT_TYPES:
            flash("Please select a valid account type.", "error")
            return render_template(
                "add_account.html", account_types=ACCOUNT_TYPES, currencies=CURRENCIES,
                form=request.form,
            )

        if currency not in CURRENCIES:
            flash("Please select a valid currency.", "error")
            return render_template(
                "add_account.html", account_types=ACCOUNT_TYPES, currencies=CURRENCIES,
                form=request.form,
            )

        try:
            opening_balance = float(opening_balance_raw) if opening_balance_raw else 0.0
        except ValueError:
            flash("Opening balance must be a number.", "error")
            return render_template(
                "add_account.html", account_types=ACCOUNT_TYPES, currencies=CURRENCIES,
                form=request.form,
            )

        insert_account(
            session["user_id"], name, account_type, currency, opening_balance, description
        )
        flash("Account added.", "success")
        return redirect(url_for("accounts"))

    return render_template(
        "add_account.html", account_types=ACCOUNT_TYPES, currencies=CURRENCIES, form={}
    )


@app.route("/accounts/<int:id>/edit", methods=["GET", "POST"])
def edit_account(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    account = get_account_by_id(id, session["user_id"])
    if account is None:
        abort(404)

    if request.method == "GET":
        return render_template(
            "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
            currencies=CURRENCIES, form={},
        )

    name = request.form.get("name", "").strip()
    account_type = request.form.get("account_type", "").strip()
    currency = request.form.get("currency", "").strip()
    opening_balance_raw = request.form.get("opening_balance", "").strip()
    description = request.form.get("description", "").strip()
    is_active = bool(request.form.get("is_active"))

    if not name:
        flash("Please enter an account name.", "error")
        return render_template(
            "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
            currencies=CURRENCIES, form=request.form,
        )

    if account_type not in ACCOUNT_TYPES:
        flash("Please select a valid account type.", "error")
        return render_template(
            "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
            currencies=CURRENCIES, form=request.form,
        )

    if currency not in CURRENCIES:
        flash("Please select a valid currency.", "error")
        return render_template(
            "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
            currencies=CURRENCIES, form=request.form,
        )

    try:
        opening_balance = float(opening_balance_raw) if opening_balance_raw else 0.0
    except ValueError:
        flash("Opening balance must be a number.", "error")
        return render_template(
            "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
            currencies=CURRENCIES, form=request.form,
        )

    if account["is_active"] and not is_active:
        active_accounts = get_accounts_by_user(session["user_id"], active_only=True)
        if len(active_accounts) <= 1:
            flash("You must keep at least one active account.", "error")
            return render_template(
                "edit_account.html", account=account, account_types=ACCOUNT_TYPES,
                currencies=CURRENCIES, form=request.form,
            )

    update_account(
        id, session["user_id"], name, account_type, currency, opening_balance,
        description, is_active,
    )
    flash("Account updated.", "success")
    return redirect(url_for("accounts"))


@app.route("/accounts/<int:id>/delete", methods=["POST"])
def delete_account(id):
    if not session.get("user_id"):
        return redirect(url_for("login"))

    account = get_account_by_id(id, session["user_id"])
    if account is None:
        abort(404)

    if account["is_active"]:
        active_accounts = get_accounts_by_user(session["user_id"], active_only=True)
        if len(active_accounts) <= 1:
            flash("You must keep at least one active account.", "error")
            return redirect(url_for("accounts"))

    result = delete_account_by_id(id, session["user_id"])
    if result == "deactivated":
        flash("Account has existing transactions, so it was deactivated instead of deleted.", "success")
    else:
        flash("Account deleted.", "success")
    return redirect(url_for("accounts"))


# Export routes respond with JSON errors instead of redirecting to /login,
# since they're called via fetch() from the download buttons, not page navigation.


@app.route("/export/monthly")
def export_monthly():
    if not session.get("user_id"):
        return jsonify({"error": "Please sign in to export your data."}), 401

    try:
        year = int(request.args.get("year", ""))
        month = int(request.args.get("month", ""))
        if not (1 <= month <= 12):
            raise ValueError
    except ValueError:
        return jsonify({"error": "Please select a valid month and year."}), 400

    date_from = date(year, month, 1).isoformat()
    date_to = date(year, month, calendar.monthrange(year, month)[1]).isoformat()

    transactions = get_all_transactions(
        session["user_id"], date_from=date_from, date_to=date_to
    )
    workbook = _build_transactions_workbook(transactions)
    filename = "expenses_income_{:04d}_{:02d}.xlsx".format(year, month)
    return _send_workbook(workbook, filename)


@app.route("/export/full")
def export_full():
    if not session.get("user_id"):
        return jsonify({"error": "Please sign in to export your data."}), 401

    transactions = get_all_transactions(session["user_id"])
    workbook = _build_transactions_workbook(transactions)
    return _send_workbook(workbook, "expenses_income_full.xlsx")


if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(debug=debug, port=int(os.environ.get("PORT", 5001)))
