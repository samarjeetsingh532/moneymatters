"""
Tests for the Excel export feature (Add Excel export)

Covers:
- GET /export/monthly and /export/full auth guard (unauthenticated -> 401 JSON, not a page redirect)
- GET /export/monthly validation (missing/non-numeric/out-of-range month/year -> 400 JSON)
- GET /export/monthly returns a real .xlsx file scoped to the selected month, with the
  documented filename pattern, containing both expense and income rows
- GET /export/full returns a real .xlsx file with the user's entire history, with the
  documented filename pattern
- Data isolation: one user's export never contains another user's transactions
- Workbook contents: header row, and Date/Amount cells use real date/numeric types
"""

import io
from datetime import date

import openpyxl
import pytest
from werkzeug.security import generate_password_hash

import database.db as db_module
from app import app as flask_app
from database.db import init_db
from database.queries import insert_expense, insert_income

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def db_path(tmp_path):
    return str(tmp_path / "test_spendly.db")


@pytest.fixture
def app(db_path, monkeypatch):
    monkeypatch.setattr(db_module, "DB_PATH", db_path)

    flask_app.config.update(
        {
            "TESTING": True,
            "SECRET_KEY": "test-secret",
            "WTF_CSRF_ENABLED": False,
        }
    )

    with flask_app.app_context():
        init_db()
        yield flask_app


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def registered_user(client):
    email = "testuser@example.com"
    password = "testpass123"
    client.post(
        "/register",
        data={
            "name": "Test User",
            "email": email,
            "password": password,
            "confirm_password": password,
        },
        follow_redirects=True,
    )
    conn = db_module.get_db()
    row = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()
    return row["id"], email, password


@pytest.fixture
def auth_client(client, registered_user):
    user_id, _email, _password = registered_user
    with client.session_transaction() as sess:
        sess["user_id"] = user_id
        sess["user_name"] = "Test User"
    return client


def _create_user(name, email):
    conn = db_module.get_db()
    cursor = conn.execute(
        "INSERT INTO users (name, email, password_hash) VALUES (?, ?, ?)",
        (name, email, generate_password_hash("pass")),
    )
    conn.commit()
    user_id = cursor.lastrowid
    conn.close()
    return user_id


def _load_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.data))


def _filename(response):
    disposition = response.headers.get("Content-Disposition", "")
    if "filename=" not in disposition:
        return None
    return disposition.split("filename=")[1].strip('"; ')


# ---------------------------------------------------------------------------
# Auth guard
# ---------------------------------------------------------------------------


class TestAuthGuard:
    def test_monthly_export_unauthenticated_returns_401_json(self, client):
        response = client.get("/export/monthly?year=2026&month=4")
        assert response.status_code == 401
        assert response.get_json()["error"]

    def test_full_export_unauthenticated_returns_401_json(self, client):
        response = client.get("/export/full")
        assert response.status_code == 401
        assert response.get_json()["error"]


# ---------------------------------------------------------------------------
# Monthly export validation
# ---------------------------------------------------------------------------


class TestMonthlyValidation:
    def test_missing_params_returns_400(self, auth_client):
        response = auth_client.get("/export/monthly")
        assert response.status_code == 400
        assert response.get_json()["error"]

    def test_non_numeric_month_returns_400(self, auth_client):
        response = auth_client.get("/export/monthly?year=2026&month=abc")
        assert response.status_code == 400

    def test_out_of_range_month_returns_400(self, auth_client):
        response = auth_client.get("/export/monthly?year=2026&month=13")
        assert response.status_code == 400

    def test_zero_month_returns_400(self, auth_client):
        response = auth_client.get("/export/monthly?year=2026&month=0")
        assert response.status_code == 400


# ---------------------------------------------------------------------------
# Monthly export happy path
# ---------------------------------------------------------------------------


class TestMonthlyExport:
    def test_returns_xlsx_with_expected_filename_and_mimetype(self, registered_user):
        user_id, _email, _password = registered_user
        insert_expense(user_id, 100.0, "Food", "2026-08-05", "Groceries")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/monthly?year=2026&month=8")

        assert response.status_code == 200
        assert response.mimetype == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert _filename(response) == "expenses_income_2026_08.xlsx"

    def test_contains_only_selected_month_transactions_of_both_types(
        self, registered_user
    ):
        user_id, _email, _password = registered_user
        insert_expense(user_id, 100.0, "Food", "2026-08-05", "Groceries")
        insert_income(user_id, 50000.0, "Salary", "2026-08-01", "August salary")
        insert_expense(user_id, 999.0, "Bills", "2026-07-15", "Not in August")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/monthly?year=2026&month=8")

        workbook = _load_workbook(response)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 2
        types = {row[3] for row in rows}
        assert types == {"Expense", "Income"}
        assert all(row[0].year == 2026 and row[0].month == 8 for row in rows)


# ---------------------------------------------------------------------------
# Full export
# ---------------------------------------------------------------------------


class TestFullExport:
    def test_returns_xlsx_with_expected_filename(self, auth_client, registered_user):
        response = auth_client.get("/export/full")
        assert response.status_code == 200
        assert _filename(response) == "expenses_income_full.xlsx"

    def test_contains_all_transactions_across_months(self, registered_user):
        user_id, _email, _password = registered_user
        insert_expense(user_id, 100.0, "Food", "2026-08-05", "Groceries")
        insert_income(user_id, 50000.0, "Salary", "2026-08-01", "August salary")
        insert_expense(user_id, 999.0, "Bills", "2026-01-15", "January bill")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/full")

        workbook = _load_workbook(response)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3

    def test_header_row_present(self, auth_client, registered_user):
        response = auth_client.get("/export/full")
        workbook = _load_workbook(response)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        assert header == ["Date", "Description", "Category", "Type", "Amount"]


# ---------------------------------------------------------------------------
# Data isolation
# ---------------------------------------------------------------------------


class TestDataIsolation:
    def test_full_export_excludes_other_users_transactions(self, registered_user):
        user_id, _email, _password = registered_user
        other_user_id = _create_user("Other User", "other@example.com")

        insert_expense(user_id, 100.0, "Food", "2026-08-05", "Mine")
        insert_expense(other_user_id, 500.0, "Shopping", "2026-08-06", "Not mine")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/full")

        workbook = _load_workbook(response)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        descriptions = [row[1] for row in rows]

        assert descriptions == ["Mine"]

    def test_monthly_export_excludes_other_users_transactions(self, registered_user):
        user_id, _email, _password = registered_user
        other_user_id = _create_user("Other User", "other2@example.com")

        insert_expense(user_id, 100.0, "Food", "2026-08-05", "Mine")
        insert_expense(other_user_id, 500.0, "Shopping", "2026-08-06", "Not mine")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/monthly?year=2026&month=8")

        workbook = _load_workbook(response)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        descriptions = [row[1] for row in rows]

        assert descriptions == ["Mine"]


# ---------------------------------------------------------------------------
# Cell data types
# ---------------------------------------------------------------------------


class TestCellTypes:
    def test_date_and_amount_use_real_data_types(self, registered_user):
        user_id, _email, _password = registered_user
        insert_expense(user_id, 123.45, "Food", "2026-08-05", "Groceries")

        with flask_app.test_client() as c:
            with c.session_transaction() as sess:
                sess["user_id"] = user_id
            response = c.get("/export/monthly?year=2026&month=8")

        workbook = _load_workbook(response)
        sheet = workbook.active
        data_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        date_value, _desc, _cat, _type, amount_value = data_row
        assert isinstance(date_value, date)
        assert isinstance(amount_value, (int, float))
        assert amount_value == 123.45
